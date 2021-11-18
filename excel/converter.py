import os
from loguru import logger
import openpyxl
from db.db import user_collection
from excel.config import list_name


def excel_to_mongo(path):
    wb = openpyxl.load_workbook(filename=path)
    sheet = wb[list_name]
    i = 2
    while True:
        id = sheet.cell(row=i, column=1).value
        name = sheet.cell(row=i, column=2).value
        stage = sheet.cell(row=i, column=3).value
        if not id:
            break
        i += 1
        print(id, name, stage)
        element = {
            "name": name,
            "stage": stage
        }
        user_collection.insert_one(element)
