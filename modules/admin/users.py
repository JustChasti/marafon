from os import remove
from datetime import date
from db.db import user_collection
from config import bot
from config import start_date
from modules.keyboards import keyboard_admin
import openpyxl
from modules.excel import converter


def add_to_user(message, data, admin_panel):
    try:
        name, mark = data.split('/')
        user = user_collection.find_one({'name': name})
        try:
            data = user['critical'] + int(mark)
            element = {
                "$set": {
                    'critical': data
                }
            }
            user_collection.update_one({'_id': user["_id"]}, element)
        except Exception as e:
            element = {
                "$set": {
                    'critical': int(mark)
                }
            }
            user_collection.update_one({'_id': user["_id"]}, element)
        bot.send_message(
            message.from_user.id,
            'Добавлено',
            reply_markup=keyboard_admin
        )
    except Exception as e:
        logger.exception(e)
        bot.send_message(
            message.from_user.id,
            str(e),
            reply_markup=keyboard_admin
        )
    bot.register_next_step_handler(message, admin_panel)


excel_path = 'modules/excel/data/stats0.xlsx'


def get_beginer_week_table(message, data, admin_panel):
    try:
        delta = date.today() - start_date
        delta = int(delta.days)
        if delta < 7:
            week = 'week 1'
        elif delta < 14:
            week = 'week 2'
        elif delta < 21:
            week = 'week 3'
        else:
            week = 'week 4'
            users = user_collection.find({'programm': 'beginer'})
            row = 2
            wb = openpyxl.load_workbook(filename=excel_path)
            sheet = wb.active
            sheet.title = 'Статистика'
            sheet[f'A1'] = 'Имя'
            sheet[f'B1'] = 'Шичко'
            sheet[f'C1'] = 'Планирование'
            sheet[f'D1'] = 'Благодарности'
            sheet[f'E1'] = 'Основные'
            sheet[f'F1'] = 'Книга'
            sheet[f'G1'] = 'Аудио книга'
            sheet[f'H1'] = 'Спорт'
            sheet[f'I1'] = 'Фитнес игра'
            sheet[f'J1'] = 'Пробежки прогулки'
            sheet[f'K1'] = 'Контрастный душ'
            sheet[f'L1'] = 'Чистые дни'
            sheet[f'N1'] = 'Прямой эфир'
            sheet[f'N1'] = 'Общий балл'
            for i in users:
                try:
                    data = i[week]
                    sum = 0
                    # sheet[f'A{row}'] = i["name"]
                    # column = 'B'
                    for j in data:
                        sum += int(data[j])
                        # sheet[f'{column}{row}'] = data[j]
                        # column = chr(ord(column) + 1)
                    # sheet[f'{column}{row}'] = sum
                    try:
                        sum += i['critical']
                    except Exception as e:
                        pass
                    sheet[f'N{row}'] = sum
                    sheet[f'A{row}'] = i['name']
                    try:
                        sheet[f'B{row}'] = data['shichko']
                    except Exception as e:
                        pass
                    try:
                        sheet[f'C{row}'] = data['planning']
                    except Exception as e:
                        pass
                    try:
                        sheet[f'D{row}'] = data['thanks']
                    except Exception as e:
                        pass
                    try:
                        sheet[f'E{row}'] = data['main_tasks']
                    except Exception as e:
                        pass
                    try:
                        sheet[f'F{row}'] = data['book']
                    except Exception as e:
                        pass
                    try:
                        sheet[f'G{row}'] = data['audio_book']
                    except Exception as e:
                        pass
                    try:
                        sheet[f'H{row}'] = data['sport']
                    except Exception as e:
                        pass
                    try:
                        sheet[f'I{row}'] = data['fitness-game']
                    except Exception as e:
                        pass
                    try:
                        sheet[f'J{row}'] = data['walk']
                    except Exception as e:
                        pass
                    try:
                        sheet[f'K{row}'] = data['shower']
                    except Exception as e:
                        pass
                    try:
                        sheet[f'L{row}'] = data['clean_day']
                    except Exception as e:
                        pass
                    try:
                        sheet[f'M{row}'] = data['stream']
                    except Exception as e:
                        pass
                    row += 1
                except Exception as e:
                    pass
            wb.save('modules/excel/data/begins_stat.xlsx')

        doc = open('modules/excel/data/begins_stat.xlsx', 'rb')
        bot.send_document(
            message.from_user.id,
            doc,
            reply_markup=keyboard_admin
        )
        doc.close()
        remove('modules/excel/data/begins_stat.xlsx')
    except Exception as e:
        logger.exception(e)
        bot.send_message(
            message.from_user.id,
            str(e),
            reply_markup=keyboard_admin
        )
    bot.register_next_step_handler(message, admin_panel)


class table_does_not_exist(Exception):
    pass


def get_potok_table(message, data, admin_panel):
    try:
        if data == 'Старт':
            programm = 'start'
        elif data == 'Лидер':
            programm = 'leader'
        elif data == 'Эксперт':
            programm = 'profi'
        else:
            raise table_does_not_exist('Такой таблицы нет')

        users = user_collection.find({'programm': programm})
        row = 2
        wb = openpyxl.load_workbook(filename=excel_path)
        sheet = wb.active
        sheet.title = 'Статистика'
        sheet[f'A1'] = 'Имя'
        sheet[f'B1'] = 'Этап'
        sheet[f'C1'] = 'Шичко'
        sheet[f'D1'] = 'Планирование'
        sheet[f'E1'] = 'Благодарности'
        sheet[f'F1'] = 'Основные'
        sheet[f'G1'] = 'Книга'
        sheet[f'H1'] = 'Аудио книга'
        sheet[f'I1'] = 'Спорт'
        sheet[f'J1'] = 'Фитнес игра'
        sheet[f'K1'] = 'Пробежки прогулки'
        sheet[f'L1'] = 'Контрастный душ'
        sheet[f'M1'] = 'Чистые дни'
        sheet[f'N1'] = 'Прямой эфир'
        sheet[f'O1'] = 'Добавочные баллы'
        sheet[f'P1'] = 'Общий балл'
        for i in users:
            sum = 0
            sheet[f'A{row}'] = i["name"]
            sheet[f'B{row}'] = i["stage"]
            try:
                sum += int(i['shichko'])
                sheet[f'C{row}'] = i['shichko']
            except Exception as e:
                sheet[f'C{row}'] = 0
            try:
                sum += int(i['plans'])
                sheet[f'D{row}'] = i['plans']
            except Exception as e:
                sheet[f'D{row}'] = 0
            try:
                sum += int(i['thanks'])
                sheet[f'E{row}'] = i['thanks']
            except Exception as e:
                sheet[f'E{row}'] = 0
            try:
                sum += int(i['main_task'])
                sheet[f'F{row}'] = i['main_task']
            except Exception as e:
                sheet[f'F{row}'] = 0
            try:
                sum += int(i['book'])
                sheet[f'G{row}'] = i['book']
            except Exception as e:
                sheet[f'G{row}'] = 0
            try:
                sum += int(i['audio_book'])
                sheet[f'H{row}'] = i['audio_book']
            except Exception as e:
                sheet[f'H{row}'] = 0
            try:
                sum += int(i['sport'])
                sheet[f'I{row}'] = i['sport']
            except Exception as e:
                sheet[f'I{row}'] = 0
            try:
                sum += int(i['fitness-game'])
                sheet[f'J{row}'] = i['fitness-game']
            except Exception as e:
                sheet[f'J{row}'] = 0
            try:
                sum += int(i['run-walk'])
                sheet[f'K{row}'] = i['run-walk']
            except Exception as e:
                sheet[f'K{row}'] = 0
            try:
                sum += int(i['shower'])
                sheet[f'L{row}'] = i['shower']
            except Exception as e:
                sheet[f'L{row}'] = 0
            try:
                sum += int(i['clean'])
                sheet[f'M{row}'] = i['clean']
            except Exception as e:
                sheet[f'M{row}'] = 0
            try:
                sum += int(i['stream'])
                sheet[f'N{row}'] = i['stream']
            except Exception as e:
                sheet[f'N{row}'] = 0
            try:
                sum += int(i['critical'])
                sheet[f'O{row}'] = i['critical']
            except Exception as e:
                sheet[f'O{row}'] = 0
            sheet[f'P{row}'] = sum
            row += 1
        wb.save('modules/excel/data/begins_stat.xlsx')

        doc = open('modules/excel/data/begins_stat.xlsx', 'rb')
        bot.send_document(
            message.from_user.id,
            doc,
            reply_markup=keyboard_admin
        )
        doc.close()
        remove('modules/excel/data/begins_stat.xlsx')
    except Exception as e:
        logger.exception(e)
        bot.send_message(
            message.from_user.id,
            str(e),
            reply_markup=keyboard_admin
        )
    bot.register_next_step_handler(message, admin_panel)


def get_all_table(message, data, admin_panel):
    try:
        users = user_collection.find({})
        row = 2
        wb = openpyxl.load_workbook(filename=excel_path)
        sheet = wb.active
        sheet.title = 'Статистика'
        sheet[f'A1'] = 'Имя'
        sheet[f'B1'] = 'Этап'
        sheet[f'C1'] = 'Шичко'
        sheet[f'D1'] = 'Планирование'
        sheet[f'E1'] = 'Благодарности'
        sheet[f'F1'] = 'Основные'
        sheet[f'G1'] = 'Книга'
        sheet[f'H1'] = 'Аудио книга'
        sheet[f'I1'] = 'Спорт'
        sheet[f'J1'] = 'Фитнес игра'
        sheet[f'K1'] = 'Пробежки прогулки'
        sheet[f'L1'] = 'Контрастный душ'
        sheet[f'M1'] = 'Чистые дни'
        sheet[f'N1'] = 'Прямой эфир'
        sheet[f'O1'] = 'Добавочные баллы'
        sheet[f'P1'] = 'Общий балл'
        for i in users:
            sum = 0
            try:
                if i['programm'] == 'beginer':
                    shichko_b = 0
                    plan_b = 0
                    thank_b = 0
                    main_b = 0
                    book_b = 0
                    audio_b = 0
                    sport_b = 0
                    fitgame_b = 0
                    run_b = 0
                    shower_b = 0
                    clean_b = 0
                    stream_b = 0
                    try:
                        sum += i['critical']
                        sheet[f'O{row}'] = i['critical']
                    except Exception as e:
                        sheet[f'O{row}'] = 0
                    for j in i:
                        if 'week' in j:
                            data = i[j]
                            for k in data:
                                sum += int(data[k])
                            try:
                                shichko_b += data['shichko']
                            except Exception as e:
                                pass
                            try:
                                plan_b += data['planning']
                            except Exception as e:
                                pass
                            try:
                                thank_b += data['thanks']
                            except Exception as e:
                                pass
                            try:
                                main_b += data['main_tasks']
                            except Exception as e:
                                pass
                            try:
                                book_b += data['book']
                            except Exception as e:
                                pass
                            try:
                                audio_b += data['audio_book']
                            except Exception as e:
                                pass
                            try:
                                sport_b += data['sport']
                            except Exception as e:
                                pass
                            try:
                                fitgame_b += data['fitness-game']
                            except Exception as e:
                                pass
                            try:
                                run_b += data['walk']
                            except Exception as e:
                                pass
                            try:
                                shower_b += data['shower']
                            except Exception as e:
                                pass
                            try:
                                clean_b += data['clean_day']
                            except Exception as e:
                                pass
                            try:
                                stream_b += data['stream']
                            except Exception as e:
                                pass
                    # вставка в таблицу:
                    sheet[f'C{row}'] = shichko_b
                    sheet[f'D{row}'] = plan_b
                    sheet[f'E{row}'] = thank_b
                    sheet[f'F{row}'] = main_b
                    sheet[f'G{row}'] = book_b
                    sheet[f'H{row}'] = audio_b
                    sheet[f'I{row}'] = sport_b
                    sheet[f'J{row}'] = fitgame_b
                    sheet[f'K{row}'] = run_b
                    sheet[f'L{row}'] = shower_b
                    sheet[f'M{row}'] = clean_b
                    sheet[f'N{row}'] = stream_b
                    sheet[f'P{row}'] = sum
                else:
                    try:
                        sum += int(i['shichko'])
                        sheet[f'C{row}'] = i['shichko']
                    except Exception as e:
                        sheet[f'C{row}'] = 0
                    try:
                        sum += int(i['plans'])
                        sheet[f'D{row}'] = i['plans']
                    except Exception as e:
                        sheet[f'D{row}'] = 0
                    try:
                        sum += int(i['thanks'])
                        sheet[f'E{row}'] = i['thanks']
                    except Exception as e:
                        sheet[f'E{row}'] = 0
                    try:
                        sum += int(i['main_task'])
                        sheet[f'F{row}'] = i['main_task']
                    except Exception as e:
                        sheet[f'F{row}'] = 0
                    try:
                        sum += int(i['book'])
                        sheet[f'G{row}'] = i['book']
                    except Exception as e:
                        sheet[f'G{row}'] = 0
                    try:
                        sum += int(i['audio_book'])
                        sheet[f'H{row}'] = i['audio_book']
                    except Exception as e:
                        sheet[f'H{row}'] = 0
                    try:
                        sum += int(i['sport'])
                        sheet[f'I{row}'] = i['sport']
                    except Exception as e:
                        sheet[f'I{row}'] = 0
                    try:
                        sum += int(i['fitness-game'])
                        sheet[f'J{row}'] = i['fitness-game']
                    except Exception as e:
                        sheet[f'J{row}'] = 0
                    try:
                        sum += int(i['run-walk'])
                        sheet[f'K{row}'] = i['run-walk']
                    except Exception as e:
                        sheet[f'K{row}'] = 0
                    try:
                        sum += int(i['shower'])
                        sheet[f'L{row}'] = i['shower']
                    except Exception as e:
                        sheet[f'L{row}'] = 0
                    try:
                        sum += int(i['clean'])
                        sheet[f'M{row}'] = i['clean']
                    except Exception as e:
                        sheet[f'M{row}'] = 0
                    try:
                        sum += int(i['stream'])
                        sheet[f'N{row}'] = i['stream']
                    except Exception as e:
                        sheet[f'N{row}'] = 0
                    try:
                        sum += int(i['critical'])
                        sheet[f'O{row}'] = i['critical']
                    except Exception as e:
                        sheet[f'O{row}'] = 0
                    sheet[f'P{row}'] = sum

            except Exception as e:
                logger.info('user not registred')

            sheet[f'A{row}'] = i['name']
            sheet[f'B{row}'] = i['stage']
            row += 1

        wb.save('modules/excel/data/begins_stat.xlsx')
        doc = open('modules/excel/data/begins_stat.xlsx', 'rb')
        bot.send_document(
            message.from_user.id,
            doc,
            reply_markup=keyboard_admin
        )
        doc.close()
        remove('modules/excel/data/begins_stat.xlsx')
    except Exception as e:
        logger.exception(e)
        bot.send_message(
            message.from_user.id,
            str(e),
            reply_markup=keyboard_admin
        )
    bot.register_next_step_handler(message, admin_panel)


def up_potok_main(message, admin_panel):
    try:
        fileID = message.document.file_id
        file_info = bot.get_file(fileID)
        data = bot.download_file(file_info.file_path)
        name = f'modules/reminder/data/{message.document.file_name}'
        with open(name, 'wb') as out:
            out.write(data)
        bot.send_message(
            message.from_user.id,
            f'загружен файл {name}',
            reply_markup=keyboard_admin
        )
    except Exception as e:
        logger.exception(e)
        bot.send_message(
            message.from_user.id,
            str(e),
            reply_markup=keyboard_admin
        )
    bot.register_next_step_handler(message, admin_panel)


def up_to_excel(message, admin_panel):
    try:
        fileID = message.document.file_id
        file_info = bot.get_file(fileID)
        data = bot.download_file(file_info.file_path)
        name = f'modules/excel/data/{message.document.file_name}'
        with open(name, 'wb') as out:
            out.write(data)
        converter.excel_to_mongo(name)
        bot.send_message(
            message.from_user.id,
            f'загружен файл {name}',
            reply_markup=keyboard_admin
        )
    except Exception as e:
        logger.exception(e)
        bot.send_message(
            message.from_user.id,
            str(e),
            reply_markup=keyboard_admin
        )
    bot.register_next_step_handler(message, admin_panel)


def send_to_potok(message, data, admin_panel):
    try:
        if data == 'Новичок':
            programm = 'beginer'
            users = user_collection.find({'programm': programm})
        elif data == 'Старт':
            programm = 'start'
            users = user_collection.find({'programm': programm})
        elif data == 'Лидер':
            programm = 'leader'
            users = user_collection.find({'programm': programm})
        elif data == 'Эксперт':
            programm = 'profi'
            users = user_collection.find({'programm': programm})
        elif data == 'Все':
            users = user_collection.find({})
        else:
            raise table_does_not_exist('Такой таблицы нет')
        for i in users:
            try:
                bot.send_message(
                    i["telegram_id"],
                    message.text,
                )
            except Exception as e:
                logger.exception(e)
                bot.send_message(
                    message.from_user.id,
                    f'Пользователь {i["name"]} не зарегистрирован или вы отправляете не текст',
                    reply_markup=keyboard_admin
                )
    except Exception as e:
        logger.exception(e)
        bot.send_message(
            message.from_user.id,
            str(e),
            reply_markup=keyboard_admin
        )
    bot.send_message(
        message.from_user.id,
        message.text,
        reply_markup=keyboard_admin
    )
    bot.register_next_step_handler(message, admin_panel)
